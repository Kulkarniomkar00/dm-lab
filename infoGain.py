import openpyxl
import math


path = "C:\\Users\\it\\Desktop\\InfoGain.xlsx"

wb_obj = openpyxl.load_workbook(path)

sheet_obj = wb_obj.active

cell_obj = sheet_obj.cell(row =1, column=2)
print(cell_obj.value)

Name_col = []

for i in range(6):
	cell_obj = sheet_obj.cell(row = 1, column= i)
	Name_Col.append(cell_obj.value)
# for col in sheet_obj["Name"]:
# 	Name_col.append(col.value)

print(Name_col)


